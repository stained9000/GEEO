from django.shortcuts import render, get_object_or_404, redirect
from .models import Escuela, Personal, Matricula, Actividad, Destreza, Presupuesto, Visita, Empleado, Municipio, Propuesta, Ofrecimiento, CodigosDE, PurchaseOrder, Servicio
from .forms import EscuelaForm, PersonalForm, MatriculaForm, DestrezaForm, ActividadForm, VisitaForm, PresupuestoForm, PropuestaForm, OfrecimientoForm, PurchaseOrderForm, FacturaForm
from django.utils import timezone
from sodapy import Socrata
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime



from io import BytesIO
from reportlab.pdfgen import canvas
from django.http import HttpResponse, HttpResponseRedirect
from reportlab.lib.pagesizes import letter, landscape
import time
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

import openpyxl
from django.core import serializers
import json

from django.views.decorators.csrf import csrf_exempt

from django.contrib.auth import authenticate as auth
from django.contrib.auth import login

# Create your views here.



def rol_check(user):
    try:
        return user.empleado.rol == 'Administrativo'
    except:
        return False

@user_passes_test(rol_check, login_url='/')
def lista_escuelas(request):

    escuelas = Escuela.objects.all()
    total_escuelas = len(escuelas)


    return render(request, 'CATEDRA/lista_escuelas.html', {'escuelas': escuelas, 'total_escuelas': total_escuelas})

@login_required
def inicio(request):
    try:
        empleado = Empleado.objects.get(usuario=request.user)
        if empleado.rol == 'Vendedor':
            return redirect('perfil_vendedor', pk=empleado.pk)
        else:
            return render(request, 'CATEDRA/inicio.html')
    except:
        return render(request, 'CATEDRA/inicio.html')

def perfil_escuela(request, codigo):

    escuela = get_object_or_404(Escuela, codigo=codigo)

    try:
        vendedor = Empleado.objects.filter(municipio__nombre__contains=escuela.municipio_escolar)[0]
    except:
        vendedor = 'Esta escuela no tiene vendedor asignado.'

    try:
        personal = Personal.objects.get(escuela=escuela)
    except:
        personal = 'Aún no se ha agregado información de personal administrativo a esta escuela.'

    try:
        matricula = Matricula.objects.get(escuela=escuela)

        total_maestros = matricula.maestros_de_espanol + matricula.maestros_de_matematicas + matricula.maestros_de_ingles + matricula.maestros_de_ciencias + matricula.maestros_otros
    except:
        matricula = 'Aún no se ha agregado información de matrícula a esta escuela.'

        total_maestros = None

    try:
        destrezas = Destreza.objects.get(escuela=escuela)
    except:
        destrezas = 'Aún no se ha agregado información de destrezas a esta escuela.'


    visitas = Visita.objects.filter(escuela=escuela).order_by("-fecha")

    if len(visitas) == 0:
        visitas = 'Aun no se ha visitado esta escuela.'


    return render(request, 'CATEDRA/perfil_escuela.html', {'escuela': escuela, 'personal': personal, 'matricula': matricula, 'total_maestros': total_maestros, 'destrezas': destrezas, 'vendedor': vendedor, 'visitas': visitas})

@login_required
def crear_escuela(request):
    if request.method == "POST":
        form = EscuelaForm(request.POST)
        if form.is_valid():
            escuela = form.save(commit=False)
            escuela.author = request.user
            escuela.fecha_creacion = timezone.now()
            escuela.save()
            return redirect('perfil_escuela', codigo=escuela.codigo)
    else:
        form = EscuelaForm()

    return render(request, 'CATEDRA/escuela_edit.html', {'form': form})

@login_required
def escuela_edit(request, codigo):
    escuela = get_object_or_404(Escuela, codigo=codigo)
    if request.method == "POST":
        form = EscuelaForm(request.POST, instance=escuela)
        if form.is_valid():
            escuela = form.save(commit=False)
            escuela.author = request.user
            escuela.fecha_creacion = timezone.now()
            escuela.save()
            return redirect('perfil_escuela', codigo=escuela.codigo)
    else:
        form = EscuelaForm(instance=escuela)

    return render(request, 'CATEDRA/escuela_edit.html', {'form': form})

@login_required
def crear_personal(request, codigo):
    escuela = Escuela.objects.get(codigo=codigo)
    if request.method == "POST":
        form = PersonalForm(request.POST)
        if form.is_valid():
            personal = form.save(commit=False)
            personal.escuela = escuela
            personal.fecha_modificacion = timezone.now()
            personal.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = PersonalForm()

    return render(request, 'CATEDRA/personal_edit.html', {'form': form, 'escuela': escuela})

@login_required
def personal_edit(request, codigo):
    escuela = Escuela.objects.get(codigo=codigo)
    personal = get_object_or_404(Personal, escuela=escuela)
    if request.method == "POST":
        form = PersonalForm(request.POST, instance=personal)
        if form.is_valid():
            personal = form.save(commit=False)
            personal.fecha_modificacion = timezone.now()
            personal.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = PersonalForm(instance=personal)

    return render(request, 'CATEDRA/personal_edit.html', {'form': form, 'escuela': escuela})

@login_required
def crear_matricula(request, codigo):
    if request.method == "POST":
        form = MatriculaForm(request.POST)
        if form.is_valid():
            matricula = form.save(commit=False)
            matricula.escuela = Escuela.objects.get(codigo=codigo)
            matricula.fecha_modificacion = timezone.now()
            matricula.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = MatriculaForm()

    return render(request, 'CATEDRA/matricula_edit.html', {'form': form})

@login_required
def matricula_edit(request, codigo):
    matricula = get_object_or_404(Matricula, escuela=Escuela.objects.get(codigo=codigo))
    if request.method == "POST":
        form = MatriculaForm(request.POST, instance=matricula)
        if form.is_valid():
            matricula = form.save(commit=False)
            matricula.fecha_modificacion = timezone.now()
            matricula.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = MatriculaForm(instance=matricula)

    return render(request, 'CATEDRA/matricula_edit.html', {'form': form})

@login_required
def crear_destrezas(request, codigo):
    escuela = Escuela.objects.get(codigo=codigo)
    if request.method == "POST":
        form = DestrezaForm(request.POST)
        if form.is_valid():
            destrezas = form.save(commit=False)
            destrezas.escuela = escuela
            destrezas.fecha_modificacion = timezone.now()
            destrezas.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = DestrezaForm()

    return render(request, 'CATEDRA/destrezas_edit.html', {'form': form, 'escuela': escuela})

@login_required
def destrezas_edit(request, codigo):
    escuela = Escuela.objects.get(codigo=codigo)
    destrezas = get_object_or_404(Destreza, escuela=Escuela.objects.get(codigo=codigo))
    if request.method == "POST":
        form = DestrezaForm(request.POST, instance=destrezas)
        if form.is_valid():
            destrezas = form.save(commit=False)
            destrezas.fecha_modificacion = timezone.now()
            destrezas.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = MatriculaForm(instance=destrezas)

    return render(request, 'CATEDRA/destrezas_edit.html', {'form': form, 'escuela': escuela})

@login_required
def crear_actividad(request, codigo):
    if request.method == "POST":
        form = ActividadForm(request.POST)
        if form.is_valid():
            actividad = form.save(commit=False)
            actividad.escuela = Escuela.objects.get(codigo=codigo)
            actividad.usuario = request.user
            actividad.fecha_creacion = timezone.now()
            actividad.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = ActividadForm()

    return render(request, 'CATEDRA/actividad_edit.html', {'form': form})

@login_required
def actividad_edit(request, codigo, pk):
    actividad = get_object_or_404(Actividad, pk=pk)
    if request.method == "POST":
        form = ActividadForm(request.POST, instance=actividad)
        if form.is_valid():
            actividad = form.save(commit=False)
            actividad.usuario = request.user
            actividad.fecha_modificacion = timezone.now()
            actividad.save()
            return redirect('perfil_escuela', codigo=codigo)
    else:
        form = ActividadForm(instance=actividad)

    return render(request, 'CATEDRA/actividad_edit.html', {'form': form})

@login_required
def cargar_escuelas(request):
    client = Socrata("data.pr.gov", None)

    # Example authenticated client (needed for non-public datasets):
    # client = Socrata(data.pr.gov,
    #                  MyAppToken,
    #                  userame="user@example.com",
    #                  password="AFakePassword")

    # First 2000 results, returned as JSON from API / converted to Python list of
    # dictionaries by sodapy.
    results = client.get("ceib-pqg3", limit=2000)

    # Convert to pandas DataFrame
    # results_df = pd.DataFrame.from_records(result_list)


    for escuela in results[900:]:
        #Filtro de codigos de escuelas invalidos
        try:
            int(escuela['codigo'])
        except:
            continue

        #Creacion o update de escuelas
        escuela_obj, escuela_created = Escuela.objects.update_or_create(
                codigo = escuela['codigo'],
                defaults={
                'nombre': escuela['escuela'],
                'telefono': '(787) ' + escuela.get('telefono', '000-0000'),
                'fax' : '(787) ' + escuela.get('fax', '000-0000'),
                'region_educativa' : escuela['region'],
                'distrito_escolar' : escuela['distrito'],
                'municipio_escolar' : escuela.get('municipio_escolar'),
                'direccion_fisica' : escuela.get('direccio_fisica'),
                'direccion_zipcode' : escuela.get('direccion_zipcode'),
                'zona' : escuela.get('zona', "RURAL"),
                'operacional' : escuela['estatus_actual'],
                'nivel' : escuela.get('nivel_original', 'Elemental'),
                'grados' : escuela.get('grados_posterior'),
                }
                )

        if escuela_created:
            escuela_obj.fecha_creacion = timezone.now()
            escuela_obj.fecha_modificacion = timezone.now()
            escuela_obj.save()
        else:
            escuela_obj.fecha_modificacion = timezone.now()
            escuela_obj.save()


        #Creacion o update de personal de las Escuelas
        #La data solo contiene informacion del director. La otra debe ser llenada manual.

        personal_obj, personal_created = Personal.objects.update_or_create(
               escuela = escuela_obj,
               defaults={
               'nombre_del_director' : escuela.get('director'),
               }
               )

        if personal_created:
            personal_obj.fecha_creacion = timezone.now()
            personal_obj.fecha_modificacion = timezone.now()
            personal_obj.save()
        else:
            personal_obj.fecha_modificacion = timezone.now()
            personal_obj.save()


        #Creacion o update de matricula de la escuela

        matricula_obj, matricula_created = Matricula.objects.update_or_create(
               escuela = escuela_obj,
               defaults={
               'matricula_de_estudiantes' : escuela.get('matricula_total'),
               'grado_infant' : escuela.get('grado_infant'),
               'grado_toddler' : escuela.get('grado_toddler'),
               'grado_pk' : escuela.get('grado_pk'),
               'grado_pke' : escuela.get('grado_pke'),
               'grado_pkm' : escuela.get('grado_pkm'),
               'grado_k' : escuela.get('grado_k'),
               'grado_1' : escuela.get('grado_1'),
               'grado_2' : escuela.get('grado_2'),
               'grado_3' : escuela.get('grado_3'),
               'grado_4' : escuela.get('grado_4'),
               'grado_5' : escuela.get('grado_5'),
               'grado_6' : escuela.get('grado_6'),
               'grado_SGE' : escuela.get('grado_SGE'),
               'grado_7' : escuela.get('grado_7'),
               'grado_8' : escuela.get('grado_8'),
               'grado_9' : escuela.get('grado_9'),
               'grado_SGI' : escuela.get('grado_SGI'),
               'grado_10' : escuela.get('grado_10'),
               'grado_11' : escuela.get('grado_11'),
               'grado_12' : escuela.get('grado_12'),
               'grado_SGS' : escuela.get('grado_SGS'),
               }
               )

        if matricula_created:
            matricula_obj.fecha_creacion = timezone.now()
            matricula_obj.fecha_modificacion = timezone.now()
            matricula_obj.save()
        else:
            matricula_obj.fecha_modificacion = timezone.now()
            matricula_obj.save()


        #Creacion o update de destrezas

        destreza_obj, destreza_created = Destreza.objects.update_or_create(
               escuela = escuela_obj,
               defaults={
               'espanol_prebasico' : escuela.get('espa_ol_pre_b_sico'),
               'espanol_basico' : escuela.get('espa_ol_b_sico'),
               'espanol_proficiente' : escuela.get('espa_ol_proficiente'),
               'espanol_avanzado' : escuela.get('espa_ol_avanzado'),
               'matematicas_prebasico' : escuela.get('matem_ticas_pre_b_sico'),
               'matematicas_basico' : escuela.get('matem_ticas_b_sico'),
               'matematicas_proficiente' : escuela.get('matem_ticas_proficiente'),
               'matematicas_avanzado' : escuela.get('matem_ticas_avanzado'),
               'ingles_prebasico' : escuela.get('ingl_s_pre_b_sico'),
               'ingles_basico' : escuela.get('ingl_s_b_sico'),
               'ingles_proficiente' : escuela.get('ingl_s_proficiente'),
               'ingles_avanzado' : escuela.get('ingl_s_avanzado'),
               }
               )

        if destreza_obj:
            destreza_obj.fecha_creacion = timezone.now()
            destreza_obj.fecha_modificacion = timezone.now()
            destreza_obj.save()
        else:
            destreza_obj.fecha_modificacion = timezone.now()
            destreza_obj.save()



    return render(request, 'CATEDRA/cargar_escuelas.html', {})

@login_required
def lista_vendedores(request):
    lista_vendedores = Empleado.objects.filter(rol='Vendedor')
    page = request.GET.get('page', 1)

    paginator = Paginator(lista_vendedores, 50)

    try:
        vendedores = paginator.page(page)
    except PageNotAnInteger:
        vendedores = paginator.page(1)
    except EmptyPage:
        vendedores = paginator.page(paginator.num_pages)

    return render(request, 'CATEDRA/lista_vendedores.html', {'vendedores': vendedores})

@login_required
def perfil_vendedor(request, pk):

    mes = datetime.now().month
    vendedor = get_object_or_404(Empleado, pk=pk)
    visitas = Visita.objects.filter(usuario=vendedor.usuario).order_by("-fecha")[0:5]
    municipios = vendedor.municipio.all()
    municipios_lista_nombres = [municipio.nombre for municipio in municipios]
    lista_escuelas = Escuela.objects.filter(municipio_escolar__in=municipios_lista_nombres)
    total_escuelas = len(lista_escuelas)
    propuestas = Propuesta.objects.filter(vendedor=vendedor).order_by("-fecha")

    lista_escuelas_visitadas = []
    lista_escuelas_sin_visita = []

    for escuela in lista_escuelas:
        if len(escuela.visita_set.filter(fecha__month=mes)) > 0:
            lista_escuelas_visitadas.append(escuela)
        else:
            lista_escuelas_sin_visita.append(escuela)

    total_escuelas_visitadas = len(lista_escuelas_visitadas)
    total_escuelas_sin_visita = len(lista_escuelas_sin_visita)

    ventas_totales = 0
    for propuesta in propuestas:
        ofrecimientos = propuesta.ofrecimiento_set.filter(estado='APROBADA')
        for ofrecimiento in ofrecimientos:
            if ofrecimiento.codigode.codigo == 11829:
                ventas_totales += ofrecimiento.codigode.costo * ofrecimiento.participantes
            else:
                ventas_totales += ofrecimiento.codigode.costo

    ventas_totales_percent = ventas_totales / 8000

    if ventas_totales_percent > 1:
        ventas_totales_percent = 1

    #paginator escuelas
    page = request.GET.get('page1', 1)

    paginator = Paginator(lista_escuelas, 6)

    try:
        escuelas = paginator.page(page)
    except PageNotAnInteger:
        escuelas = paginator.page(1)
    except EmptyPage:
        escuelas = paginator.page(paginator.num_pages)

    #paginator escuelas visitadas
    page = request.GET.get('page2', 1)

    paginator = Paginator(lista_escuelas_visitadas, 6)

    try:
        lista_escuelas_visitadas = paginator.page(page)
    except PageNotAnInteger:
        lista_escuelas_visitadas = paginator.page(1)
    except EmptyPage:
        lista_escuelas_visitadas = paginator.page(paginator.num_pages)

    return render(request, 'CATEDRA/perfil_vendedor.html', {'vendedor': vendedor, 'visitas': visitas, 'municipios': municipios, 'total_escuelas': total_escuelas, 'propuestas': propuestas, 'ventas_totales': ventas_totales, 'ventas_totales_percent': ventas_totales_percent , 'lista_escuelas_visitadas': lista_escuelas_visitadas, 'lista_escuelas_sin_visita': lista_escuelas_sin_visita, 'total_escuelas_visitadas': total_escuelas_visitadas, 'total_escuelas_sin_visita': total_escuelas_sin_visita,})


@login_required
def crear_visita(request, pk_vendedor):
    next = request.GET.get('next', '/')
    empleado = Empleado.objects.get(pk=pk_vendedor)
    municipios = empleado.municipio.all()
    lista_municipios = [ municipio.nombre for municipio in municipios ]

    if request.method == "POST":
        form = VisitaForm(request.POST, request.FILES)
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')
        if form.is_valid():
            visita = form.save(commit=False)
            visita.usuario = empleado.usuario
            visita.fecha_creacion = timezone.now()
            visita.save()
            return  HttpResponseRedirect(next)
    else:
        form = VisitaForm()
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')

    return render(request, 'CATEDRA/visita_edit.html', {'form': form})

@login_required
def visita_edit(request, pk_vendedor, pk_visita):
    next = request.GET.get('next', '/')
    empleado = Empleado.objects.get(pk=pk_vendedor)
    municipios = empleado.municipio.all()
    lista_municipios = [ municipio.nombre for municipio in municipios ]
    visita = get_object_or_404(Visita, pk=pk_visita)
    if request.method == "POST":
        form = VisitaForm(request.POST, instance=visita)
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')
        if form.is_valid():
            visita = form.save(commit=False)
            visita.fecha_modificacion = timezone.now()
            visita.save()
            return HttpResponseRedirect(next)
    else:
        form = VisitaForm(instance=visita)
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')

    return render(request, 'CATEDRA/visita_edit.html', {'form': form})

@login_required
def historial_visitas(request):
    if request.user.empleado.rol == 'Vendedor':
        vendedor = get_object_or_404(Empleado, pk=request.user.empleado.pk)
        visitas = Visita.objects.filter(usuario=vendedor.usuario).order_by("-fecha")
    else:
        vendedor = 'Usuario no es vendedor'
        visitas = Visita.objects.all()

    return render(request, 'CATEDRA/historial_visitas.html', {'vendedor': vendedor, 'visitas': visitas,})

@login_required
def historial_propuestas(request, pk):
    vendedor = get_object_or_404(Empleado, pk=pk)
    if vendedor.rol == 'Vendedor':
        propuestas = Propuesta.objects.filter(vendedor=vendedor).order_by("-fecha")
    else:
        propuestas = Propuesta.objects.all()

    return render(request, 'CATEDRA/historial_propuestas.html', {'vendedor': vendedor, 'propuestas': propuestas,})

@login_required
def crear_propuesta(request, pk_vendedor):
    empleado = Empleado.objects.get(pk=pk_vendedor)
    municipios = empleado.municipio.all()
    lista_municipios = [ municipio.nombre for municipio in municipios ]

    if request.method == "POST":
        form = PropuestaForm(request.POST, request.FILES)

        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')
        if form.is_valid():
            propuesta = form.save(commit=False)
            propuesta.vendedor = empleado
            propuesta.save()
            return redirect('propuesta_detalle', pk_propuesta=propuesta.pk)
    else:
        form = PropuestaForm()
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')

    return render(request, 'CATEDRA/propuesta_edit.html', {'form': form})

@login_required
def propuesta_edit(request, pk_vendedor, pk_propuesta):
    empleado = Empleado.objects.get(pk=pk_vendedor)
    municipios = empleado.municipio.all()
    lista_municipios = [ municipio.nombre for municipio in municipios ]
    propuesta = get_object_or_404(Propuesta, pk=pk_propuesta)

    if request.method == "POST":
        form = PropuestaForm(request.POST, instance=propuesta)
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')
        if form.is_valid():
            propuesta = form.save(commit=False)
            propuesta.save()
            return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)
    else:
        form = PropuestaForm(instance=propuesta)
        form.fields['escuela'].queryset = Escuela.objects.filter(municipio_escolar__in=lista_municipios).order_by('nombre')

    return render(request, 'CATEDRA/propuesta_edit.html', {'form': form})

@login_required
def propuesta_detalle(request, pk_propuesta):
    propuesta = Propuesta.objects.get(pk=pk_propuesta)
    personal = Personal.objects.get(escuela=propuesta.escuela)
    ofrecimientos = Ofrecimiento.objects.filter(propuesta=propuesta)
    pos = PurchaseOrder.objects.filter(propuesta=propuesta)

    costo_total = 0

    for ofrecimiento in ofrecimientos:
        if ofrecimiento.codigode.codigo == 11829:
            costo_total += ofrecimiento.codigode.costo * ofrecimiento.participantes
        else:
            costo_total += ofrecimiento.codigode.costo

    return render(request, 'CATEDRA/propuesta_detalle.html', {'propuesta': propuesta, 'personal': personal, 'ofrecimientos': ofrecimientos, 'costo_total': costo_total, 'pos': pos,})

@login_required
def crear_ofrecimiento(request, pk_propuesta):
    request_type = request.path[-4:]
    propuesta = Propuesta.objects.get(pk=pk_propuesta)
    codigosde = CodigosDE.objects.filter(tipo=propuesta.tipo).order_by('codigo')
    estrategias = Servicio.objects.all().order_by('estrategia').values('estrategia').distinct()
    titulos = Servicio.objects.all().order_by('titulo').values('titulo').distinct()

    if request.method == "POST":
        form = OfrecimientoForm(request.POST, request.FILES)
        if form.is_valid():
            ofrecimiento = form.save(commit=False)
            ofrecimiento.propuesta = Propuesta.objects.get(pk=pk_propuesta)
            ofrecimiento.save()
            return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)
    else:
        form = OfrecimientoForm()

    return render(request, 'CATEDRA/ofrecimiento_edit.html', {'form': form, 'codigosde': codigosde, 'estrategias': estrategias, 'titulos': titulos, 'request_type': request_type,})

@login_required
def ofrecimiento_edit(request, pk_ofrecimiento, pk_propuesta):
    request_type = request.path[-4:]
    propuesta = Propuesta.objects.get(pk=pk_propuesta)
    codigosde = CodigosDE.objects.filter(tipo=propuesta.tipo).order_by('codigo')
    estrategias = Servicio.objects.all().order_by('estrategia').values('estrategia').distinct()
    titulos = Servicio.objects.all().order_by('titulo').values('titulo').distinct()
    print(request.path)

    ofrecimiento = get_object_or_404(Ofrecimiento, pk=pk_ofrecimiento)

    if request.method == "POST":
        form = OfrecimientoForm(request.POST, instance=ofrecimiento)
        if form.is_valid():
            ofrecimiento = form.save(commit=False)
            ofrecimiento.save()
            return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)
    else:
        form = OfrecimientoForm(instance=ofrecimiento)


    return render(request, 'CATEDRA/ofrecimiento_edit.html', {'form': form, 'codigosde': codigosde, 'estrategias': estrategias, 'titulos': titulos, 'request_type': request_type,})

@login_required
def borrar_ofrecimiento(request, pk_propuesta, pk_ofrecimiento):
    ofrecimiento = Ofrecimiento.objects.get(pk=pk_ofrecimiento)
    ofrecimiento.delete()

    return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)

@login_required
def propuesta_pdf(request, pk_propuesta):
    from reportlab.lib.units import inch

    propuesta = Propuesta.objects.get(pk=pk_propuesta)
    personal = Personal.objects.get(escuela=propuesta.escuela)
    ofrecimientos = Ofrecimiento.objects.filter(propuesta=propuesta)

    def xstr(s):
        if s is None:
            return ''
        else:
            return str(s)

    def format_thousands(value):
        if value == None:
            return None
        else:
            try:
                return "$ {:,}".format(value)
            except:
                return None

     # Create the HttpResponse object with the appropriate PDF headers.
    filename = 'attachment; filename=' + '"Porpuesta No.' + str(propuesta.pk) + ' - ' + str(propuesta.escuela.nombre) + '.pdf"'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = filename

    buffer = BytesIO()

    # Create the PDF object, using the BytesIO object as its "file."
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = letter
    # Draw things on the PDF. Here's where the PDF generation happens.
    # See the ReportLab documentation for the full list of functionality.
    image = ImageReader('https://www.geeopr.com/index_htm_files/63.png')
    p.setLineWidth(.3)
    p.setFont('Helvetica', 12)
    p.translate(inch,inch*7.5)

    #Titulo
    if propuesta.tipo == 'Publica - Maestros':
        p.drawImage(image, 0, -30, width=100, height=100, mask='auto')
        p.drawString(144, 0, 'PROPUESTA DE TALLERES DE ADECUACION PROFESIONAL')
    elif propuesta.tipo == 'Colegio - Maestros':
        p.drawImage(image, 0, -30, width=100, height=100, mask='auto')
        p.drawString(144, 0, 'PROPUESTA DE TALLERES DE ADECUACION PROFESIONAL COLEGIOS')
    elif propuesta.tipo == 'Publica - Padres':
        p.drawImage(image, 0, -30, width=100, height=100, mask='auto')
        p.drawString(144, 0, 'PROPUESTA DE TALLERES A PADRES')
    else:
        p.drawImage(image, 0, -30, width=100, height=100, mask='auto')
        p.drawString(144, 0, 'PROPUESTA DE TALLERES A PADRES COLEGIOS')


    #Linea 1
    p.setFont('Helvetica', 10)
    p.drawString(0,-26,'Escuela: ' + propuesta.escuela.nombre)
    p.line(40, -28, 203, -28)

    p.drawString(216,-26,'Código: ' + xstr(propuesta.escuela.codigo))
    p.line(254,-28,417,-28)

    p.drawString(430, -26, 'Tel/Fax: ' + propuesta.escuela.telefono)
    p.line(470, -28, 650, -28)

    #Linea 2
    p.drawString(0, -42, 'Región Educativa: ' + propuesta.escuela.region_educativa)
    p.line(84, -44, 203, -44)

    p.drawString(216, -42, 'Distrito: ' + propuesta.escuela.distrito_escolar)
    p.line(254, -44, 417, -44)

    p.drawString(430, -42, 'Dirección: ' + propuesta.escuela.direccion_fisica)
    p.line(477, -44, 650, -44)

    #Linea 3
    p.drawString(0, -58, "Director(a): " + xstr(personal.nombre_del_director))
    p.line(53, -60, 203, -60)

    p.drawString(216, -58, 'E-mail: ' + xstr(personal.email_del_director))
    p.line(250, -60, 417, -60)

    #table
    costo_total = 0
    data= [['Cantidad', 'Materia', 'Modalidad/Codigo', 'Estrategia', 'Titulo', 'Horas', 'Costo', 'Total'],
           ]
    long_title = False
    for ofrecimiento in ofrecimientos:
        if ofrecimiento.codigode.codigo == 11829:
            costo_total += ofrecimiento.codigode.costo * ofrecimiento.participantes
            data.append([str(ofrecimiento.participantes), ofrecimiento.materia, str(ofrecimiento.codigode.modalidad) + " / " + str(ofrecimiento.codigode.codigo), ofrecimiento.estrategia, ofrecimiento.titulo, str(ofrecimiento.horas), str(format_thousands(ofrecimiento.codigode.costo)), str(format_thousands(ofrecimiento.codigode.costo*ofrecimiento.participantes))])
            if len(ofrecimiento.titulo) >= 156:
                long_title = True
        else:
            costo_total += ofrecimiento.codigode.costo
            data.append(["1", ofrecimiento.materia, str(ofrecimiento.codigode.modalidad) + " / " + str(ofrecimiento.codigode.codigo), ofrecimiento.estrategia, ofrecimiento.titulo, str(ofrecimiento.horas), str(format_thousands(ofrecimiento.codigode.costo)), str(format_thousands(ofrecimiento.codigode.costo))])
            if len(ofrecimiento.titulo) >= 156:
                long_title = True


    if propuesta.tipo == 'Colegio - Maestros':
        data.append(['Total de Maestros:', '', str(ofrecimientos[0].participantes), '', '', '', 'Total', str(format_thousands(costo_total)),])
        data.append([''' <u>Necesidad</u>: ''' + str(propuesta.necesidad), '', '', '', '<u>Descripción</u>: ' + str(propuesta.descripcion), '', '', ''])
    elif propuesta.tipo == 'Colegio - Padres':
        data.append(['', '', '', '', '', '', 'Total', str(format_thousands(costo_total)),])
        data.append([''' <u>Necesidad</u>: ''' + str(propuesta.necesidad), '', '', '', '<u>Descripción</u>: ' + str(propuesta.descripcion), '', '', ''])
    else:
        data.append(['', '', '', '', '', '', 'Total', str(format_thousands(costo_total)),])

    s = getSampleStyleSheet()
    s = s["BodyText"]
    s.wordWrap = 'CJK'
    data2 = [[Paragraph(cell, s) for cell in row] for row in data]
    if long_title == True:
        t=Table(data2, colWidths=[inch*0.8, inch*0.8, inch*0.9, inch*1, inch*3, inch*0.6, inch*1, inch*1],  rowHeights=inch*0.8)
    else:
        t=Table(data2, colWidths=[inch*0.8, inch*0.8, inch*0.9, inch*1, inch*3, inch*0.6, inch*1, inch*1],  rowHeights=inch*0.7)

    if propuesta.tipo == 'Colegio - Maestros':
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,1),'CENTER'),
                            ('VALIGN',(0,0),(-1,-2),'MIDDLE'),
                            ('SPAN',(0,-2),(1,-2)),
                            ('SPAN',(3,-2),(5,-2)),
                            ('SPAN',(0,-1),(3,-1)),
                            ('SPAN',(-4,-1),(-1,-1)),
                            ('VALIGN',(0,-1),(-1,-1),'TOP'),
                            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                            ]))
    elif propuesta.tipo == 'Colegio - Padres':
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,1),'CENTER'),
                            ('VALIGN',(0,0),(-1,-2),'MIDDLE'),
                            ('SPAN',(0,-2),(5,-2)),
                            ('SPAN',(0,-1),(3,-1)),
                            ('SPAN',(-4,-1),(-1,-1)),
                            ('VALIGN',(0,-1),(-1,-1),'TOP'),
                            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                            ]))
    else:
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,1),'CENTER'),
                            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                            ('SPAN',(0,-1),(-3,-1)),
                            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                            ]))
    if long_title == True:
        tableheight = len(data) * inch * 0.8 + 76
    else:
        tableheight = len(data) * inch * 0.7 + 76

    t.wrap(648, 358)
    t.drawOn(p, 0, -tableheight)

    #


    #Firma
    p.drawString(0, -435, 'Firma: ')
    p.line(30, -437, 180, -437)

    p.drawString(216, -435, 'Fecha Servicio: ')
    p.line(287,-437,394,-437)

    #Footer
    p.line(0, -455, 650, -455)
    p.line(0, -455, 0, -530)
    p.line(0, -530, 650, -530)
    p.line(650, -455, 650, -530)

    p.drawString(260, -473, 'Representante de Servicio')
    p.drawString(5, -491, 'Compañia: Global Education Exchange Opportunities, Inc.')
    p.line(55, -493, 265, -493)

    p.drawString(342, -491, 'SS Patronal: 660723113')
    p.line(400, -493, 453, -493)

    p.drawString(530, -491, 'E-mail: info@geeopr.com')
    p.line(564, -493, 644, -493)

    p.drawString(5, -509, 'Nombre Representante: ' + propuesta.vendedor.usuario.first_name + " " + propuesta.vendedor.usuario.last_name)
    p.line(115, -511, 265, -511)

    p.drawString(342, -509, 'Telefono: ' + propuesta.vendedor.telefono)
    p.line(387, -511, 455, -511)

    # Close the PDF object cleanly.
    p.showPage()
    p.save()

    # Get the value of the BytesIO buffer and write it to the response.
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def lista_ofrecimientos(request):

    if request.user.empleado.rol == 'Administrativo':
        ofrecimientos = Ofrecimiento.objects.all()
    else:
        ofrecimientos = Ofrecimiento.objects.filter(propuesta__vendedor__usuario=request.user)

    return render(request, 'CATEDRA/lista_ofrecimientos.html', {'ofrecimientos': ofrecimientos})


@login_required
def crear_po(request, pk_propuesta):
    propuesta = Propuesta.objects.get(pk=pk_propuesta)

    if request.method == "POST":
        form = PurchaseOrderForm(request.POST, request.FILES)
        form.fields['ofrecimiento'].queryset = Ofrecimiento.objects.filter(propuesta=propuesta, estado='EN EVALUACION')
        if form.is_valid():
            purchase_order = form.save(commit=False)
            purchase_order.propuesta = Propuesta.objects.get(pk=pk_propuesta)
            purchase_order.save()
            form.save_m2m()
            po = PurchaseOrder.objects.get(numero=purchase_order.numero)
            ofrecimientos = po.ofrecimiento.all()

            for ofrecimiento in ofrecimientos:
                ofrecimiento.estado = 'APROBADA'
                ofrecimiento.save()
            return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)
    else:
        form = PurchaseOrderForm()
        form.fields['ofrecimiento'].queryset = Ofrecimiento.objects.filter(propuesta=propuesta, estado='EN EVALUACION')

    return render(request, 'CATEDRA/po_edit.html', {'form': form})

@login_required
def borrar_po(request, pk_propuesta, pk_po):
    po = PurchaseOrder.objects.get(pk=pk_po)
    ofrecimientos = po.ofrecimiento.all()

    for ofrecimiento in ofrecimientos:
        ofrecimiento.estado = 'EN EVALUACION'
        ofrecimiento.save()

    po.delete()

    return redirect('propuesta_detalle', pk_propuesta=pk_propuesta)

@login_required
def po_edit(request, pk_po, pk_propuesta):
    next = request.GET.get('next', '/')
    po = get_object_or_404(PurchaseOrder, pk=pk_po)
    propuesta = Propuesta.objects.get(pk=pk_propuesta)
    ofrecimiento_queryset = po.ofrecimiento.all() | Ofrecimiento.objects.filter(propuesta=propuesta, estado='EN EVALUACION')

    for ofrecimiento in ofrecimiento_queryset:
        ofrecimiento.estado = 'EN EVALUACION'
        ofrecimiento.save()



    if request.method == "POST":
        form = PurchaseOrderForm(request.POST, instance=po)
        form.fields['ofrecimiento'].queryset = ofrecimiento_queryset
        if form.is_valid():
            purchase_order = form.save(commit=False)
            purchase_order.save()
            form.save_m2m()
            ofrecimientos = po.ofrecimiento.all()

            for ofrecimiento in ofrecimientos:
                ofrecimiento.estado = 'APROBADA'
                ofrecimiento.save()

            return HttpResponseRedirect(next)
    else:
        form = PurchaseOrderForm(instance=po)
        form.fields['ofrecimiento'].queryset = ofrecimiento_queryset

    return render(request, 'CATEDRA/po_edit.html', {'form': form})

@login_required
def lista_po(request):

    if request.user.empleado.rol == 'Administrativo':
        pos = PurchaseOrder.objects.all()
    else:
        pos = PurchaseOrder.objects.filter(propuesta__vendedor__usuario=request.user)

    return render(request, 'CATEDRA/lista_po.html', {'pos': pos})

@login_required
def cargar_servicios(request):
    wb = openpyxl.load_workbook('CATEDRA/catalogocatedra.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    for i in range(1, 146):
        obj, created = Servicio.objects.get_or_create(
                    tipo = sheet.cell(row=i, column=1).value,
                    estrategia = sheet.cell(row=i, column=2).value,
                    titulo = sheet.cell(row=i, column=3).value
                    )

        print(obj, created)

    sheet2 = wb.get_sheet_by_name('Sheet2')

    for i in range(1, 82):
        obj, created = Servicio.objects.get_or_create(
                    tipo = sheet2.cell(row=i, column=1).value,
                    estrategia = sheet2.cell(row=i, column=2).value,
                    titulo = sheet2.cell(row=i, column=3).value
                    )

        print(obj, created)

    return render(request, 'CATEDRA/cargar_servicios.html', {})

@login_required
def catalogo_servicios(request):
    servicios = Servicio.objects.all()

    return render(request, 'CATEDRA/catalogo_servicios.html', {'servicios': servicios,})

def ofrecimiento_titulos_json(request, estrategia):
    print(estrategia)
    titulos = Servicio.objects.filter(estrategia=estrategia).order_by('titulo').values('titulo').distinct()
    lista_titulos = [{'titulo': titulos[i]['titulo']} for i in range(0, len(titulos))]


    print(lista_titulos)
    json_titulos = json.dumps(lista_titulos)
    print(json_titulos)
    return HttpResponse(json_titulos, content_type="application/javascript")

def po_detalle(request, numero_po):
    po = PurchaseOrder.objects.get(numero=numero_po)
    ofrecimientos = po.ofrecimiento.all()
    facturas = po.factura_set.all()

    costo_total = 0

    for ofrecimiento in ofrecimientos:
        if ofrecimiento.codigode.codigo == 11829:
            costo_total += ofrecimiento.codigode.costo * ofrecimiento.participantes
        else:
            costo_total += ofrecimiento.codigode.costo

    return render(request, 'CATEDRA/PO_detalle.html', {'po': po, 'ofrecimientos': ofrecimientos, 'costo_total': costo_total, 'facturas': facturas,})

@login_required
def crear_factura(request, numero_po):
    po = PurchaseOrder.objects.get(numero=numero_po)

    if request.method == "POST":
        form = FacturaForm(request.POST)
        if form.is_valid():
            factura = form.save(commit=False)
            factura.po = po
            factura.save()
            return redirect('po_detalle', numero_po=po.numero)
    else:
        form = FacturaForm()

    return render(request, 'CATEDRA/factura_edit.html', {'form': form,})

@login_required
def factura_edit(request, numero_factura, numero_po):
    po = PurchaseOrder.objects.get(numero=numero_po)
    factura = get_object_or_404(Factura, numero=numero_factura)

    if request.method == "POST":
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            factura = form.save(commit=False)
            factura.save()
            return redirect('po_detalle', numero_po=po.numero)
    else:
        form = FacturaForm(instance=factura)


    return render(request, 'CATEDRA/factura_edit.html', {'form': form,})

@login_required
def borrar_factura(request, numero_factura, numero_po):
    factura = Factura.objects.get(numero=numero_factura)
    factura.delete()

    return redirect('po_detalle', numero_po=numero_po)
